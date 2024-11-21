import os, requests, asyncio, urllib.parse, json, re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, parse_qs
import google.generativeai as genai
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from PIL import Image
from io import BytesIO
from tqdm.asyncio import tqdm
import nodriver as uc
import absl.logging
from openpyxl.drawing.image import Image as OpenpyxlImage
from modules.services.ai import ai_parse 
from modules.utils.validators import is_valid_image
from modules.utils.extractors import extract_it_id

    
absl.logging.set_verbosity("error")

timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%Mm%Ss")

with open("url.txt", "r") as file:
    urls = file.read().splitlines()


results = {}
lock = asyncio.Lock()

async def fetch_page_source(url, folder_path):
    browser = await uc.start()

    try:
        page = await asyncio.wait_for(
            browser.get(url), timeout=10
        )  # 10초의 타임아웃 설정

        count = 0
        while True:
            if count > 5:
                async with lock:
                    results[url]["결과"] = "실패"
                return ""

            await asyncio.sleep(3)
            count += 1

            try:
                source = await asyncio.wait_for(page.get_content(), timeout=5)
            except asyncio.TimeoutError:
                async with lock:
                    results[url]["결과"] = "실패"
                return ""

            if "<img" in source:
                return source
    except asyncio.TimeoutError:
        async with lock:
            results[url]["결과"] = "실패"
        return ""
    except Exception:
        async with lock:
            results[url]["결과"] = "실패"
        return ""
    finally:
        browser.stop()


def parse_images(html_data, url):
    soup = BeautifulSoup(html_data, "html.parser")

    for tag in soup.find_all(["header", "head", "footer"]):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "recommend" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "relate" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "logo" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "together" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "list" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "review" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "banner" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "category" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "option" in class_name
    ):
        tag.decompose()
    for tag in soup.find_all(
        class_=lambda class_name: class_name and "guide" in class_name
    ):
        tag.decompose()

    img_tags = soup.find_all("img")
    img_urls = [
        (
            urljoin(url, img["src"])
            if ";base64," not in img["src"]
            else (
                urljoin(url, img["ec-data-src"]) if "ec-data-src" in img.attrs else ""
            )
        )
        for img in img_tags
        if "src" in img.attrs
        and not img["src"].lower().endswith(".svg")
        and not "//img.echosting.cafe24.com/" in img["src"]
        and "/theme/" not in img["src"]
        and "facebook" not in img["src"]
        and "icon" not in img["src"]
        and "logo" not in img["src"]
        and "common" not in img["src"]
        and "banner" not in img["src"]
        and "brand" not in img["src"]
    ]

    return img_urls


with open("config.json", "r", encoding="utf-8") as file:
    config_data = json.load(file)
    api_key = config_data["api_key"]
    model_name = config_data["model"]

genai.configure(api_key=api_key)
model = genai.GenerativeModel(
    model_name=model_name,
    generation_config={
        "temperature": 0,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 500,
        "response_mime_type": "application/json",
    },
)
model.generate_content("Hi")
os.system("cls")



async def download_images(url, folder_name):
    item_id = extract_it_id(url)

    folder_path = os.path.join(f"이미지/{item_id}", folder_name)

    os.makedirs(folder_path, exist_ok=True)

    html_data = await fetch_page_source(url, folder_path)
    img_urls = parse_images(html_data, url)

    thumb_path = ""
    if len(img_urls) > 0:
        idx = 0
        for img_url in tqdm(img_urls, desc="이미지 다운로드 중...", leave=False):
            if img_url == "":
                continue

            headers = {
                "Upgrade-Insecure-Requests": "1",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
                "sec-ch-ua": '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
            }
            try:
                img_response = requests.get(
                    img_url, headers=headers, stream=True, timeout=120
                )
                img_response.raise_for_status()
            except Exception:
                continue

            if not is_valid_image(img_response.content):
                continue

            parsed_url = urlparse(img_url)
            file_name, file_extension = os.path.splitext(
                os.path.basename(parsed_url.path)
            )

            if not file_extension:
                file_extension = ".jpg"

            img_path = os.path.join(folder_path, f"{item_id}_{idx}{file_extension}")

            with open(img_path, "wb") as file:
                for chunk in img_response.iter_content(1024):
                    file.write(chunk)

            if idx == 0:
                thumb_path = img_path

            idx += 1

        async with lock:
            results[url]["결과"] = "성공"
            results[url]["이미지"] = thumb_path

        try:
            parsed_data = ai_parse(
                ai_model=model,
                html_data=html_data)

            async with lock:
                results[url]["매장가"] = parsed_data["market_price"]
                results[url]["단가"] = parsed_data["price"]
                results[url]["성별"] = parsed_data["gender"]
                results[url]["상품명"] = (
                    re.match(r"^\[.*?\] (.*)", str(parsed_data["kor_name"])).group(1)
                    if re.match(r"^\[.*?\] (.*)", str(parsed_data["kor_name"]))
                    else str(parsed_data["kor_name"])
                )
                results[url]["영문명"] = (
                    re.match(r"^\[.*?\] (.*)", str(parsed_data["eng_name"])).group(1)
                    if re.match(r"^\[.*?\] (.*)", str(parsed_data["eng_name"]))
                    else str(parsed_data["eng_name"])
                )
                results[url]["브랜드"] = parsed_data["brand"].upper()
                results[url]["2차"] = parsed_data["first_category"]
                results[url]["3차"] = parsed_data["second_category"]
                results[url]["추가 정보\n모델명"] = str(parsed_data["genuine_number"])
                results[url]["필수옵션\n색상"] = ",".join(parsed_data["colors"])
                results[url]["필수옵션\n사이즈"] = (
                    ",".join(parsed_data["sizes"]).replace("(", "[").replace(")", "]")
                )
        except Exception:
            async with lock:
                results[url]["결과"] = "실패"
    else:
        async with lock:
            results[url]["결과"] = "실패"


async def main():
    # browser = await uc.start(headless=True)

    for url in tqdm(urls, desc="링크 순회 중..."):
        folder_name = datetime.now().strftime(timestamp)

        async with lock:
            results[urllib.parse.unquote(url)] = {
                "결과": "",
                # "링크": urllib.parse.unquote(url),
                # "상품넘버": str(folder_name),
                "상품넘버": (
                    f'=HYPERLINK("{str(urllib.parse.unquote(url))}", "{str(folder_name)}")'
                ),
                "거래처": "",
                "단가": "",
                "이미지": "",
                "1차": "",
                "2차": "",
                "3차": "",
                "4차": "",
                "필터": "",
                "성별": "",
                "브랜드": "",
                "2차 브랜드": "",
                "상품명": "",
                "영문명": "",
                "추가 정보\n모델명": "",
                "추가 정보\n배송방법": "항공특송",
                "추가 정보\n소재": "",
                "추가 정보\n구성품": "풀박스",
                "매장가": "",
                "판매가1": "",
                "판매가2": "",
                "판매가3": "",
                "필수옵션\n등급선택": "",
                "필수옵션\n사이즈": "",
                "필수옵션\n색상": "",
                "필수옵션\n굽높이": "",
                "필수옵션\n버클": "",
                "필수옵션\n도금방식": "",
                "필수옵션\n밴드": "",
            }

        await download_images(urllib.parse.unquote(url), str(folder_name))


uc.loop().run_until_complete(main())
final = list(results.values())

avail_brands = [
    "ASK YOURSELF",
    "ACNE STUDIOS",
    "ALEXANDER MCQUEEN",
    "ALEXANDER WANG",
    "ALYX",
    "AMI",
    "AMIRI",
    "ARCTERYX",
    "AUDEMARS PIGUET",
    "BALENCIAGA",
    "BALMAIN",
    "BAPE",
    "BERLUTI",
    "BLANCPAIN",
    "BOTTEGA VENETA",
    "BREGUET",
    "BALLY",
    "BREITLING",
    "BRUNELLO CUCINELLI",
    "BULGARI",
    "BURBERRY",
    "CANADA GOOSE",
    "CARTIER",
    "CASABLANCA",
    "CELINE",
    "CHANEL",
    "CHAUMET",
    "CHLOE",
    "CHROME HEARTS",
    "COMME DES GARCONS",
    "CP COMPANY",
    "DELVAUX",
    "DRIES VAN NOTEN",
    "DIESEL",
    "DIOR",
    "DOLCE & GABBANA",
    "EMPORIO ARMANI",
    "FEAR OF GOD",
    "FENDI",
    "FERRAGAMO",
    "GALLERY DEPT",
    "GENTLE MONSTER",
    "GIVENCHY",
    "GOLDEN GOOSE",
    "GOYARD",
    "GUCCI",
    "HERMES",
    "HUBLOT",
    "ISABEL MARANT",
    "IAB STUDIO",
    "IWC",
    "JACQUEMUS",
    "JIL SANDER",
    "JUNJI",
    "JIMMY CHOO",
    "JORDAN",
    "JUNYA WATANABE",
    "KENZO",
    "LANVIN BLANC",
    "LANVIN",
    "LEMAIRE",
    "LOEWE",
    "LORO PIANA",
    "LOUBOUTIN",
    "LOUIS VUITTON",
    "MACKAGE",
    "MAISON MARGIELA",
    "MAISON MIHARA YASUHIRO",
    "MANOLO BLAHNIK",
    "MARNI",
    "MARTINE ROSE",
    "MAX MARA",
    "MAISON KITSUNE",
    "MIU MIU",
    "MONCLER",
    "MOOSE KNUCKLES",
    "NEW BALANCE",
    "NIKE",
    "OFF WHITE",
    "OMEGA",
    "PHILIPP PLEIN",
    "PANERAI",
    "PARAJUMPERS",
    "PALM ANGELS",
    "PALACE",
    "PATEK PHILIPPE",
    "PRADA",
    "PIAGET",
    "POLORALPHLAUREN",
    "RAY BAN",
    "RHUDE",
    "RICK OWENS",
    "RIMOWA",
    "ROGER VIVIER",
    "ROLEX",
    "SACAI",
    "SUPREME",
    "SAINT LAURENT",
    "SALOMON",
    "STUSSY",
    "STONE ISLAND",
    "TAG HEUER",
    "THE NORTH FACE",
    "THOM BROWNE",
    "TIFFANY & CO",
    "TOM FORD",
    "TUDOR",
    "UMA WANG",
    "VACHERON CONSTANTIN",
    "VALENTINO",
    "VETEMENTS",
    "VANCLEEF",
    "VERSACE",
    "WOOYOUNGMI",
    "YEEZY",
    "ZEGNA",
    "OTHERS",
]

avail_1st_categories = [
    "상의",
    "아우터",
    "하의",
    "가방",
    "신발",
    "지갑",
    "시계",
    "패션잡화",
    "액세서리",
    "벨트",
]

avail_2nd_categories = [
    "반팔 티셔츠",
    "긴팔 티셔츠",
    "니트/가디건",
    "맨투맨",
    "후드",
    "원피스",
    "셔츠",
    "드레스",
    "슬리브리스",
    "셋업",
    "기타 상의",
    "집업",
    "자켓",
    "패딩",
    "레더",
    "코트",
    "기타 아우터",
    "팬츠",
    "쇼츠",
    "트레이닝 팬츠",
    "데님",
    "스커트",
    "기타 하의",
    "미니백",
    "백팩",
    "숄더백",
    "토트백",
    "크로스백",
    "클러치",
    "캐리어",
    "핸드백",
    "더플백",
    "버킷백",
    "기타 가방",
    "스니커즈",
    "샌들/슬리퍼",
    "플랫",
    "로퍼",
    "더비/레이스업",
    "힐/펌프스",
    "부츠",
    "기타 신발",
    "반지갑",
    "카드지갑",
    "지퍼장지갑",
    "중/장지갑",
    "여권지갑",
    "WOC",
    "기타 지갑",
    "메탈",
    "가죽",
    "우레탄",
    "머플러/스카프",
    "아이웨어",
    "넥타이",
    "모자",
    "헤어액세서리",
    "기타 잡화",
    "반지",
    "목걸이",
    "팔찌",
    "귀걸이",
    "키링",
    "브로치",
    "기타 ACC",
]

for index, data in enumerate(final):
    if data["브랜드"] not in avail_brands:
        final[index]["브랜드"] = ""
    if data["2차"] not in avail_1st_categories:
        final[index]["2차"] = ""
    if data["3차"] not in avail_2nd_categories:
        final[index]["3차"] = ""

df = pd.DataFrame(final)

wb = openpyxl.Workbook()
ws = wb.active

ws.append(list(df.columns))
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
font = Font(name="Arial")

for cell in ws[1]:
    cell.alignment = alignment
    cell.font = font

for idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
    img_path = row[4]
    if os.path.exists(img_path) and img_path != "":
        img = OpenpyxlImage(img_path)
        img.width, img.height = 80, 80
        ws.add_image(img, f"E{idx + 2}")
        ws.row_dimensions[idx + 2].height = 65
        row[4] = ""

    ws.append(row)

    for cell in ws[idx + 2]:
        cell.alignment = alignment
        cell.font = font

ws.column_dimensions["B"].width = 18
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 8.25
ws.column_dimensions["G"].width = 11.25
ws.column_dimensions["H"].width = 12.75
ws.column_dimensions["K"].width = 12.75
ws.column_dimensions["L"].width = 15
ws.column_dimensions["N"].width = 39
ws.column_dimensions["O"].width = 22.5
ws.column_dimensions["P"].width = 12
ws.column_dimensions["Q"].width = 12
ws.column_dimensions["R"].width = 12
ws.column_dimensions["S"].width = 12
ws.column_dimensions["T"].width = 12
ws.column_dimensions["U"].width = 12
ws.column_dimensions["V"].width = 12
ws.column_dimensions["W"].width = 12
ws.column_dimensions["X"].width = 12
ws.column_dimensions["Y"].width = 20
ws.column_dimensions["Z"].width = 20
ws.column_dimensions["AA"].width = 12
ws.column_dimensions["AB"].width = 12
ws.column_dimensions["AC"].width = 12
ws.column_dimensions["AD"].width = 12

wb.save(f"결과_{timestamp}.xlsx")

input("\n작업 완료! 엔터를 눌러 종료하세요 : ")
