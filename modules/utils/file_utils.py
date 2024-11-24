import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

from openpyxl.drawing.image import Image as OpenpyxlImage

def read_urls(path="url.txt"):
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            pass
    with open(path, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())



async def make_excel(final_result, timestamp):
        
    avail_brands = [
        "ASK YOURSELF",
        "ACNE STUDIOS",
        "ALEXANDER MCQUEEN",
        "ALEXANDER WANG",
        "ALYX",
        "AMI",
        "AMIRI",
        "ARCTERYX",
        "AUDEMARS PIGUET",
        "BALENCIAGA",
        "BALMAIN",
        "BAPE",
        "BERLUTI",
        "BLANCPAIN",
        "BOTTEGA VENETA",
        "BREGUET",
        "BALLY",
        "BREITLING",
        "BRUNELLO CUCINELLI",
        "BULGARI",
        "BURBERRY",
        "CANADA GOOSE",
        "CARTIER",
        "CASABLANCA",
        "CELINE",
        "CHANEL",
        "CHAUMET",
        "CHLOE",
        "CHROME HEARTS",
        "COMME DES GARCONS",
        "CP COMPANY",
        "DELVAUX",
        "DRIES VAN NOTEN",
        "DIESEL",
        "DIOR",
        "DOLCE & GABBANA",
        "EMPORIO ARMANI",
        "FEAR OF GOD",
        "FENDI",
        "FERRAGAMO",
        "GALLERY DEPT",
        "GENTLE MONSTER",
        "GIVENCHY",
        "GOLDEN GOOSE",
        "GOYARD",
        "GUCCI",
        "HERMES",
        "HUBLOT",
        "ISABEL MARANT",
        "IAB STUDIO",
        "IWC",
        "JACQUEMUS",
        "JIL SANDER",
        "JUNJI",
        "JIMMY CHOO",
        "JORDAN",
        "JUNYA WATANABE",
        "KENZO",
        "LANVIN BLANC",
        "LANVIN",
        "LEMAIRE",
        "LOEWE",
        "LORO PIANA",
        "LOUBOUTIN",
        "LOUIS VUITTON",
        "MACKAGE",
        "MAISON MARGIELA",
        "MAISON MIHARA YASUHIRO",
        "MANOLO BLAHNIK",
        "MARNI",
        "MARTINE ROSE",
        "MAX MARA",
        "MAISON KITSUNE",
        "MIU MIU",
        "MONCLER",
        "MOOSE KNUCKLES",
        "NEW BALANCE",
        "NIKE",
        "OFF WHITE",
        "OMEGA",
        "PHILIPP PLEIN",
        "PANERAI",
        "PARAJUMPERS",
        "PALM ANGELS",
        "PALACE",
        "PATEK PHILIPPE",
        "PRADA",
        "PIAGET",
        "POLORALPHLAUREN",
        "RAY BAN",
        "RHUDE",
        "RICK OWENS",
        "RIMOWA",
        "ROGER VIVIER",
        "ROLEX",
        "SACAI",
        "SUPREME",
        "SAINT LAURENT",
        "SALOMON",
        "STUSSY",
        "STONE ISLAND",
        "TAG HEUER",
        "THE NORTH FACE",
        "THOM BROWNE",
        "TIFFANY & CO",
        "TOM FORD",
        "TUDOR",
        "UMA WANG",
        "VACHERON CONSTANTIN",
        "VALENTINO",
        "VETEMENTS",
        "VANCLEEF",
        "VERSACE",
        "WOOYOUNGMI",
        "YEEZY",
        "ZEGNA",
        "OTHERS",
    ]

    avail_1st_categories = [
        "상의",
        "아우터",
        "하의",
        "가방",
        "신발",
        "지갑",
        "시계",
        "패션잡화",
        "액세서리",
        "벨트",
    ]

    avail_2nd_categories = [
        "반팔 티셔츠",
        "긴팔 티셔츠",
        "니트/가디건",
        "맨투맨",
        "후드",
        "원피스",
        "셔츠",
        "드레스",
        "슬리브리스",
        "셋업",
        "기타 상의",
        "집업",
        "자켓",
        "패딩",
        "레더",
        "코트",
        "기타 아우터",
        "팬츠",
        "쇼츠",
        "트레이닝 팬츠",
        "데님",
        "스커트",
        "기타 하의",
        "미니백",
        "백팩",
        "숄더백",
        "토트백",
        "크로스백",
        "클러치",
        "캐리어",
        "핸드백",
        "더플백",
        "버킷백",
        "기타 가방",
        "스니커즈",
        "샌들/슬리퍼",
        "플랫",
        "로퍼",
        "더비/레이스업",
        "힐/펌프스",
        "부츠",
        "기타 신발",
        "반지갑",
        "카드지갑",
        "지퍼장지갑",
        "중/장지갑",
        "여권지갑",
        "WOC",
        "기타 지갑",
        "메탈",
        "가죽",
        "우레탄",
        "머플러/스카프",
        "아이웨어",
        "넥타이",
        "모자",
        "헤어액세서리",
        "기타 잡화",
        "반지",
        "목걸이",
        "팔찌",
        "귀걸이",
        "키링",
        "브로치",
        "기타 ACC",
    ]

    for index, data in enumerate(final_result):
        if data["브랜드"] not in avail_brands:
            final_result[index]["브랜드"] = ""
        if data["2차"] not in avail_1st_categories:
            final_result[index]["2차"] = ""
        if data["3차"] not in avail_2nd_categories:
            final_result[index]["3차"] = ""

    df = pd.DataFrame(final_result)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(list(df.columns))
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font = Font(name="Arial")

    for cell in ws[1]:
        cell.alignment = alignment
        cell.font = font

    for idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        img_path = row[4]
        if os.path.exists(img_path) and img_path != "":
            img = OpenpyxlImage(img_path)
            img.width, img.height = 80, 80
            ws.add_image(img, f"E{idx + 2}")
            ws.row_dimensions[idx + 2].height = 65
            row[4] = ""

        ws.append(row)

        for cell in ws[idx + 2]:
            cell.alignment = alignment
            cell.font = font

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8.25
    ws.column_dimensions["G"].width = 11.25
    ws.column_dimensions["H"].width = 12.75
    ws.column_dimensions["K"].width = 12.75
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["N"].width = 39
    ws.column_dimensions["O"].width = 22.5
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["R"].width = 12
    ws.column_dimensions["S"].width = 12
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 12
    ws.column_dimensions["V"].width = 12
    ws.column_dimensions["W"].width = 12
    ws.column_dimensions["X"].width = 12
    ws.column_dimensions["Y"].width = 20
    ws.column_dimensions["Z"].width = 20
    ws.column_dimensions["AA"].width = 12
    ws.column_dimensions["AB"].width = 12
    ws.column_dimensions["AC"].width = 12
    ws.column_dimensions["AD"].width = 12

    wb.save(f"결과_{timestamp}.xlsx")